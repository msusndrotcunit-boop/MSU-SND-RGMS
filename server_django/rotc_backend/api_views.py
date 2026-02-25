import base64
import csv
import json
from datetime import datetime

from django.http import HttpResponse, JsonResponse, StreamingHttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from openpyxl import load_workbook
import logging
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.conf import settings
import mimetypes

from rgms.models import Cadet, MeritDemeritLog, TrainingStaff, AdminProfile

# Image compression utilities
from io import BytesIO
from PIL import Image
import os
import re

IMAGE_MAX_BYTES = int(os.getenv("IMAGE_MAX_BYTES", str(512 * 1024)))
IMAGE_MAX_DIMENSION = int(os.getenv("IMAGE_MAX_DIMENSION", str(1600)))
JPEG_QUALITY_START = int(os.getenv("JPEG_QUALITY_START", "85"))
JPEG_QUALITY_MIN = int(os.getenv("JPEG_QUALITY_MIN", "55"))
ALLOW_WEBP = os.getenv("ALLOW_WEBP", "true").lower() == "true"
PDF_MAX_BYTES = int(os.getenv("PDF_MAX_BYTES", str(2 * 1024 * 1024)))

def _downscale_preserve_aspect(img, max_dim):
    w, h = img.size
    if max(w, h) <= max_dim:
        return img
    if w >= h:
        new_w = max_dim
        new_h = int(h * (max_dim / w))
    else:
        new_h = max_dim
        new_w = int(w * (max_dim / h))
    return img.resize((new_w, new_h), Image.LANCZOS)

def _save_to_bytes(img, fmt, quality=None, progressive=False, lossless=False):
    buf = BytesIO()
    save_kwargs = {}
    if fmt.upper() == "JPEG":
        save_kwargs.update(dict(format="JPEG", quality=quality or JPEG_QUALITY_START, optimize=True, progressive=progressive))
        if img.mode in ("RGBA", "LA"):
            img = img.convert("RGB")
    elif fmt.upper() == "PNG":
        save_kwargs.update(dict(format="PNG", optimize=True))
    elif fmt.upper() == "WEBP" and ALLOW_WEBP:
        save_kwargs.update(dict(format="WEBP", quality=quality or JPEG_QUALITY_START, method=6))
        if lossless:
            save_kwargs["lossless"] = True
    else:
        raise ValueError("Unsupported format")
    img.save(buf, **save_kwargs)
    return buf.getvalue()

def compress_uploaded_image(uploaded_file, max_bytes=IMAGE_MAX_BYTES, max_dim=IMAGE_MAX_DIMENSION):
    try:
        original_bytes = uploaded_file.read()
        original_size = len(original_bytes)
        uploaded_file.seek(0)
        img = Image.open(BytesIO(original_bytes))
        img.load()
    except Exception:
        raise ValueError("Unsupported or corrupt image file")

    base_img = _downscale_preserve_aspect(img, max_dim)

    original_fmt = (img.format or "JPEG").upper()
    candidates = []
    if original_fmt == "PNG":
        if ALLOW_WEBP:
            candidates.append(("WEBP", False))
        candidates.append(("PNG", False))
        candidates.append(("JPEG", True))
    elif original_fmt == "WEBP":
        candidates.append(("WEBP", False))
        candidates.append(("JPEG", True))
    else:
        candidates.append(("JPEG", True))
        if ALLOW_WEBP:
            candidates.append(("WEBP", False))

    best_bytes = None
    best_fmt = None
    for fmt, progressive in candidates:
        if fmt == "PNG":
            try:
                data = _save_to_bytes(base_img, "PNG")
                if best_bytes is None or len(data) < len(best_bytes):
                    best_bytes = data
                    best_fmt = "PNG"
                if len(data) <= max_bytes:
                    return {
                        "bytes": data,
                        "size": len(data),
                        "format": "PNG",
                        "original_size": original_size,
                    }
            except Exception:
                pass
            continue
        for quality in range(JPEG_QUALITY_START, JPEG_QUALITY_MIN - 1, -7):
            try:
                data = _save_to_bytes(base_img, fmt, quality=quality, progressive=progressive, lossless=False)
                if best_bytes is None or len(data) < len(best_bytes):
                    best_bytes = data
                    best_fmt = fmt
                if len(data) <= max_bytes:
                    return {
                        "bytes": data,
                        "size": len(data),
                        "format": fmt,
                        "original_size": original_size,
                    }
            except Exception:
                continue

    if best_bytes is None:
        raise ValueError("Failed to compress image")
    if len(best_bytes) > max_bytes:
        return {
            "bytes": best_bytes,
            "size": len(best_bytes),
            "format": best_fmt or original_fmt,
            "original_size": original_size,
            "limit_not_met": True,
        }
    return {
        "bytes": best_bytes,
        "size": len(best_bytes),
        "format": best_fmt or original_fmt,
        "original_size": original_size,
    }

def compress_pdf_bytes(pdf_bytes, max_bytes=PDF_MAX_BYTES):
    original_size = len(pdf_bytes)
    try:
        import pikepdf
        with pikepdf.open(BytesIO(pdf_bytes)) as pdf:
            # Aggressive object stream and image recompression where possible
            out = BytesIO()
            pdf.save(out, optimize_version=True, compress_streams=True)
            data = out.getvalue()
            if len(data) <= max_bytes or len(data) < original_size:
                return {"bytes": data, "size": len(data), "original_size": original_size}
            # If still large, try linearize off to avoid extra overhead
            out2 = BytesIO()
            pdf.save(out2, linearize=True)
            data2 = out2.getvalue()
            best = data2 if len(data2) < len(data) else data
            return {"bytes": best, "size": len(best), "original_size": original_size, "limit_not_met": len(best) > max_bytes}
    except Exception as exc:
        logging.warning("PDF compression unavailable or failed: %s", exc)
        return {"bytes": pdf_bytes, "size": original_size, "original_size": original_size, "skipped": True}


def transmute_grade(value):
    if value >= 95:
        return "1.00"
    if value >= 90:
        return "1.50"
    if value >= 85:
        return "2.00"
    if value >= 80:
        return "2.50"
    if value >= 75:
        return "3.00"
    return "5.00"


def compute_cadet_grades(cadet):
    if cadet.attendance_total > 0:
        attendance_score = (cadet.attendance_present / cadet.attendance_total) * 30
    else:
        attendance_score = 0
    base_aptitude = 100 + cadet.merit_points - cadet.demerit_points
    if base_aptitude < 0:
        base_aptitude = 0
    if base_aptitude > 100:
        base_aptitude = 100
    aptitude_score = base_aptitude * 0.3
    subject_score = (
        (cadet.prelim_score + cadet.midterm_score + cadet.final_score) / 300
    ) * 40
    final_grade = attendance_score + aptitude_score + subject_score
    cadet.attendance_score = attendance_score
    cadet.aptitude_score = aptitude_score
    cadet.subject_score = subject_score
    cadet.final_grade = final_grade
    cadet.transmuted_grade = transmute_grade(final_grade)


def transparent_png_response():
    data = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR4nGNgYAAAAAMAAWgmWQ0AAAAASUVORK5CYII="
    )
    return HttpResponse(data, content_type="image/png")


def cadet_to_dict(cadet):
    return {
        "id": cadet.id,
        "student_id": cadet.student_id,
        "first_name": cadet.first_name,
        "last_name": cadet.last_name,
        "middle_name": cadet.middle_name,
        "suffix_name": cadet.suffix_name,
        "rank": cadet.rank,
        "email": cadet.email,
        "username": cadet.username,
        "contact_number": cadet.contact_number,
        "address": cadet.address,
        "gender": cadet.gender,
        "religion": cadet.religion,
        "birthdate": cadet.birthdate.isoformat() if cadet.birthdate else None,
        "course": cadet.course,
        "year_level": cadet.year_level,
        "school_year": cadet.school_year,
        "battalion": cadet.battalion,
        "company": cadet.company,
        "platoon": cadet.platoon,
        "cadet_course": cadet.cadet_course,
        "semester": cadet.semester,
        "corp_position": cadet.corp_position,
        "status": cadet.status,
        "is_profile_completed": cadet.is_profile_completed,
        "profile_pic": cadet.profile_pic.url if cadet.profile_pic else None,
        "attendance_present": cadet.attendance_present,
        "attendance_total": cadet.attendance_total,
        "attendanceScore": cadet.attendance_score,
        "merit_points": cadet.merit_points,
        "demerit_points": cadet.demerit_points,
        "aptitudeScore": cadet.aptitude_score,
        "prelim_score": cadet.prelim_score,
        "midterm_score": cadet.midterm_score,
        "final_score": cadet.final_score,
        "subjectScore": cadet.subject_score,
        "finalGrade": cadet.final_grade,
        "transmutedGrade": cadet.transmuted_grade,
        "grade_status": cadet.grade_status,
        "grade_remarks": cadet.grade_remarks,
    }


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def build_cadet_payload(data):
    def _sanitize_name(s):
        if s is None:
            return ""
        s = str(s).strip()
        # Normalize exotic apostrophes to ASCII
        s = s.replace("’", "'").replace("‘", "'").replace("`", "'").replace("“", '"').replace("”", '"')
        # Remove control characters
        s = re.sub(r"[\x00-\x1F\x7F]", "", s)
        # If only quotes or punctuation remain, treat as empty
        if re.fullmatch(r"[\'\"\\-_.\\s]*", s):
            return ""
        # Collapse multiple spaces
        s = re.sub(r"\s{2,}", " ", s)
        return s
    def _sanitize_id(s):
        if s is None:
            return ""
        return str(s).strip()
    payload = {
        "rank": _sanitize_name(data.get("rank", "")),
        "first_name": _sanitize_name(data.get("firstName") or data.get("first_name", "")),
        "middle_name": _sanitize_name(data.get("middleName") or data.get("middle_name", "")),
        "last_name": _sanitize_name(data.get("lastName") or data.get("last_name", "")),
        "suffix_name": _sanitize_name(data.get("suffixName") or data.get("suffix_name", "")),
        "student_id": _sanitize_id(data.get("studentId") or data.get("student_id", "")),
        "email": (data.get("email", "") or "").strip(),
        "username": _sanitize_id(data.get("username", "")),
        "contact_number": (data.get("contactNumber") or data.get("contact_number", "") or "").strip(),
        "address": (data.get("address", "") or "").strip(),
        "gender": _sanitize_name(data.get("gender", "")),
        "religion": _sanitize_name(data.get("religion", "")),
        "birthdate": parse_date(data.get("birthdate")),
        "course": _sanitize_name(data.get("course", "")),
        "year_level": _sanitize_name(data.get("yearLevel") or data.get("year_level", "")),
        "school_year": _sanitize_name(data.get("schoolYear") or data.get("school_year", "")),
        "battalion": _sanitize_name(data.get("battalion", "")),
        "company": _sanitize_name(data.get("company", "")),
        "platoon": _sanitize_name(data.get("platoon", "")),
        "cadet_course": _sanitize_name(data.get("cadetCourse") or data.get("cadet_course", "")),
        "semester": _sanitize_name(data.get("semester", "")),
        "corp_position": _sanitize_name(data.get("corpPosition") or data.get("corp_position", "")),
        "status": _sanitize_name(data.get("status", "Ongoing")) or "Ongoing",
    }
    # Ensure name fallbacks to avoid malformed one-char punctuation names
    if not payload["first_name"] and payload["last_name"]:
        payload["first_name"] = "Unknown"
    if not payload["last_name"] and payload["first_name"]:
        payload["last_name"] = "Cadet"
    return payload


def find_column_value(row, candidates):
    if not row:
        return None
    normalized_candidates = [
        c.strip().lower().replace(" ", "").replace("_", "").replace("-", "") for c in candidates
    ]
    for key, value in row.items():
        if value is None or value == "":
            continue
        norm_key = str(key).strip().lower().replace(" ", "").replace("_", "").replace("-", "")
        for cand in normalized_candidates:
            if norm_key == cand:
                return value
    return None


def load_rows_from_upload(uploaded_file):
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        decoded = uploaded_file.read().decode("utf-8", errors="ignore")
        reader = csv.DictReader(decoded.splitlines())
        return [dict(row) for row in reader]
    if name.endswith(".xlsx") or name.endswith(".xls"):
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None
            data_rows.append(item)
        return data_rows
    raise ValueError("Unsupported file type. Please upload CSV or Excel (.xlsx) file.")


@csrf_exempt
@require_http_methods(["GET", "PUT", "POST"])
def admin_profile_view(request):
    admin_profile, _ = AdminProfile.objects.get_or_create(username="admin")

    if request.method == "GET":
        return JsonResponse(
            {
                "id": admin_profile.id,
                "username": admin_profile.username,
                "first_name": admin_profile.first_name,
                "last_name": admin_profile.last_name,
                "email": admin_profile.email,
                "profile_pic": admin_profile.profile_pic.url if admin_profile.profile_pic else None,
                "role": "admin",
                "profile_completed": True,
            }
        )

    if request.method in ["PUT", "POST"]:
        # Handle profile picture upload
        profile_pic = request.FILES.get("profilePic")
        if profile_pic:
            allowed_types = ["image/jpeg", "image/png", "image/webp", "image/gif"]
            if profile_pic.content_type not in allowed_types:
                return JsonResponse({"message": "Invalid file type. Only JPEG, PNG, or WebP are allowed."}, status=400)

            try:
                c = compress_uploaded_image(profile_pic, max_bytes=IMAGE_MAX_BYTES, max_dim=IMAGE_MAX_DIMENSION)
            except ValueError as exc:
                return JsonResponse({"message": str(exc)}, status=400)
            if c.get("limit_not_met"):
                return JsonResponse({"message": "Unable to meet size limit after compression."}, status=400)
            from django.core.files.base import ContentFile
            content = ContentFile(c["bytes"])

            # Refined naming convention: admin_[id]_[timestamp].[ext]
            import os
            import time
            ext_map = {"JPEG": ".jpg", "JPG": ".jpg", "PNG": ".png", "WEBP": ".webp"}
            ext = ext_map.get(c["format"].upper(), os.path.splitext(profile_pic.name)[1] or ".jpg")
            filename = f"admin_{admin_profile.id}_{int(time.time())}{ext}"

            admin_profile.profile_pic = content
            admin_profile.profile_pic.name = filename
            admin_profile.save()

            return JsonResponse({
                "message": "Profile picture updated successfully",
                "profile_pic": admin_profile.profile_pic.url,
                "compression": {
                    "original": c["original_size"],
                    "final": c["size"],
                }
            })
        
        return JsonResponse({"message": "No profile picture provided"}, status=400)


@require_http_methods(["GET"])
def admin_cadets_list_view(request):
    cadets = Cadet.objects.all().order_by("last_name", "first_name")
    data = [cadet_to_dict(c) for c in cadets]
    return JsonResponse(data, safe=False)


@require_http_methods(["GET"])
def admin_cadets_archived_list_view(request):
    cadets = Cadet.objects.filter(status="Completed").order_by(
        "last_name", "first_name"
    )
    data = [cadet_to_dict(c) for c in cadets]
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["GET", "POST"])
def admin_cadets_collection_view(request):
    if request.method == "GET":
        return admin_cadets_list_view(request)
    return admin_cadets_create_view(request)


@csrf_exempt
@require_http_methods(["POST"])
def admin_cadets_create_view(request):
    if request.content_type and "application/json" in request.content_type:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    else:
        payload = request.POST
    cadet_data = build_cadet_payload(payload)
    if not cadet_data.get("student_id"):
        return JsonResponse({"message": "studentId is required"}, status=400)
    try:
        cadet, created = Cadet.objects.get_or_create(
            student_id=cadet_data["student_id"], defaults=cadet_data
        )
        if not created:
            for key, value in cadet_data.items():
                setattr(cadet, key, value)
            cadet.save()
        logging.info("Cadet %s (%s) %s", f"{cadet.last_name}, {cadet.first_name}", cadet.student_id, "created" if created else "updated")
        return JsonResponse(cadet_to_dict(cadet), status=201 if created else 200)
    except Exception as exc:
        logging.exception("Cadet creation failed: %s", exc)
        return JsonResponse({"message": "Cadet creation failed"}, status=500)


@csrf_exempt
@require_http_methods(["PUT"])
def admin_cadets_update_view(request, cadet_id):
    try:
        cadet = Cadet.objects.get(id=cadet_id)
    except Cadet.DoesNotExist:
        return JsonResponse({"message": "Cadet not found"}, status=404)
    if request.content_type and "application/json" in request.content_type:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    else:
        payload = request.POST
    cadet_data = build_cadet_payload(payload)
    for key, value in cadet_data.items():
        setattr(cadet, key, value)
    cadet.save()
    return JsonResponse(cadet_to_dict(cadet))


@csrf_exempt
@require_http_methods(["POST"])
def admin_cadets_bulk_delete_view(request):
    payload = json.loads(request.body.decode("utf-8") or "{}")
    ids = payload.get("ids") or []
    Cadet.objects.filter(id__in=ids).update(status="Completed")
    return JsonResponse({"deleted_ids": ids})


@csrf_exempt
@require_http_methods(["POST"])
def admin_cadets_bulk_restore_view(request):
    payload = json.loads(request.body.decode("utf-8") or "{}")
    ids = payload.get("ids") or []
    Cadet.objects.filter(id__in=ids).update(status="Ongoing")
    return JsonResponse({"restored_ids": ids})


@csrf_exempt
@require_http_methods(["POST"])
def admin_cadets_bulk_unlock_view(request):
    payload = json.loads(request.body.decode("utf-8") or "{}")
    ids = payload.get("ids") or []
    Cadet.objects.filter(id__in=ids).update(is_profile_completed=False)
    return JsonResponse({"unlocked_ids": ids})


@csrf_exempt
@require_http_methods(["POST"])
def admin_import_cadets_view(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        return JsonResponse({"message": "No file uploaded"}, status=400)
    try:
        rows = load_rows_from_upload(uploaded)
    except ValueError as exc:
        return JsonResponse({"message": str(exc)}, status=400)

    imported = 0
    updated = 0
    skipped = 0
    errors = []

    def _san(s):
        return (s or "").strip()
    def _san_name(s):
        return build_cadet_payload({"firstName": s}).get("first_name", "")
    for row in rows:
        try:
            custom_username = find_column_value(row, ["Username", "User Name", "username"])
            email = find_column_value(row, ["Email", "E-mail", "Email Address", "email"])
            first_name = find_column_value(row, ["First Name", "first_name", "FName", "Given Name"])
            last_name = find_column_value(row, ["Last Name", "last_name", "LName", "Surname"])
            middle_name = (
                find_column_value(row, ["Middle Name", "middle_name", "MName", "Middle Initial"]) or ""
            )
            rank = find_column_value(row, ["Rank", "rank", "Grade"]) or "Cdt"
            raw_student_id = find_column_value(
                row, ["Student ID", "student_id", "ID", "Student Number", "USN"]
            )

            if not first_name or not last_name:
                full_name = find_column_value(
                    row, ["Name", "name", "Full Name", "Cadet Name"]
                )
                if full_name:
                    parts = str(full_name).split(",")
                    if len(parts) >= 2:
                        last_name = _san_name(parts[0])
                        rest = parts[1].strip().split(" ")
                        first_name = _san_name(rest[0])
                        middle_name = _san_name(" ".join(rest[1:]) if len(rest) > 1 else middle_name)
                    else:
                        space_parts = str(full_name).split(" ")
                        if len(space_parts) >= 2:
                            last_name = _san_name(space_parts[-1])
                            first_name = _san_name(" ".join(space_parts[:-1]))
                        else:
                            first_name = _san_name(full_name)
                            last_name = last_name or "Unknown"

            student_id = _san(raw_student_id) or _san(custom_username) or _san(email)
            if not student_id and first_name:
                base = str(first_name).strip().lower()
                base = "".join(ch for ch in base if ch.isalnum())
                student_id = f"{base}{Cadet.objects.count()+1}"

            if not student_id:
                skipped += 1
                errors.append(
                    "Could not determine identity (Missing Student ID, Username, Email, or Name)"
                )
                continue

            # Sanitize final names
            sf = build_cadet_payload({
                "firstName": first_name or "Unknown",
                "lastName": last_name or "Cadet",
                "middleName": middle_name,
                "rank": rank
            })
            cadet_defaults = {
                "last_name": sf["last_name"] or "Cadet",
                "first_name": sf["first_name"] or "Unknown",
                "middle_name": sf["middle_name"],
                "suffix_name": "",
                "rank": sf["rank"] or "Cdt",
                "email": _san(email),
                "contact_number": "",
                "address": "",
                "course": "",
                "year_level": "",
                "school_year": "",
                "battalion": "",
                "company": "",
                "platoon": "",
                "cadet_course": "",
                "semester": "",
            }

            cadet, created = Cadet.objects.get_or_create(
                student_id=str(student_id), defaults=cadet_defaults
            )
            if not created:
                for key, value in cadet_defaults.items():
                    setattr(cadet, key, value)
                cadet.save()
                updated += 1
            else:
                imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(str(exc))

    return JsonResponse(
        {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10],
            "message": f"Import complete. Success: {imported + updated}, Failed: {skipped}",
        }
    )


@require_http_methods(["GET"])
def admin_notifications_list_view(request):
    return JsonResponse({"notifications": [], "unread_count": 0})


@csrf_exempt
@require_http_methods(["DELETE"])
def admin_notifications_clear_view(request):
    return JsonResponse({"cleared": True})


@require_http_methods(["GET"])
def staff_list_view(request):
    staff = TrainingStaff.objects.filter(is_archived=False).order_by("last_name", "first_name")
    data = [
        {
            "id": s.id,
            "rank": s.rank,
            "first_name": s.first_name,
            "middle_name": s.middle_name,
            "last_name": s.last_name,
            "suffix_name": s.suffix_name,
            "email": s.email,
            "contact_number": s.contact_number,
            "role": s.role,
            "profile_pic": s.profile_pic.url if s.profile_pic else None,
            "is_profile_completed": s.is_profile_completed,
            "is_archived": s.is_archived,
        }
        for s in staff
    ]
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["GET", "PUT"])
def cadet_profile_view(request):
    cadet_id_hdr = request.headers.get("X-Cadet-Id") or request.GET.get("cadetId")
    cadet = None
    if cadet_id_hdr:
        try:
            cadet = Cadet.objects.get(id=int(cadet_id_hdr))
        except (Cadet.DoesNotExist, ValueError):
            cadet = None
    if cadet is None:
        cadet = Cadet.objects.first()
    if not cadet:
        return JsonResponse({"message": "No cadet found"}, status=404)

    if request.method == "GET":
        return JsonResponse(cadet_to_dict(cadet))

    if request.method == "PUT":
        if request.content_type and "application/json" in request.content_type:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        else:
            payload = request.POST
        
        # Handle file upload if present
        profile_pic = request.FILES.get("profilePic")
        if profile_pic:
            try:
                c = compress_uploaded_image(profile_pic, max_bytes=IMAGE_MAX_BYTES, max_dim=IMAGE_MAX_DIMENSION)
            except ValueError as exc:
                return JsonResponse({"message": str(exc)}, status=400)
            if c.get("limit_not_met"):
                return JsonResponse({"message": "Unable to meet size limit after compression."}, status=400)
            from django.core.files.base import ContentFile
            content = ContentFile(c["bytes"])
            import time
            ext_map = {"JPEG": ".jpg", "JPG": ".jpg", "PNG": ".png", "WEBP": ".webp"}
            ext = ext_map.get(c["format"].upper(), ".jpg")
            content.name = f"cadet_{cadet.id}_{int(time.time())}{ext}"
            cadet.profile_pic = content

        # Update other fields
        for key in payload:
            if hasattr(cadet, key):
                setattr(cadet, key, payload[key])
        
        cadet.save()
        data = cadet_to_dict(cadet)
        if profile_pic:
            data["compression"] = {"original": c["original_size"], "final": c["size"]}
        return JsonResponse(data)


@csrf_exempt
@require_http_methods(["GET", "PUT"])
def staff_me_view(request):
    staff = TrainingStaff.objects.first()
    if not staff:
        return JsonResponse({"message": "No staff found"}, status=404)

    if request.method == "GET":
        return JsonResponse({
            "id": staff.id,
            "first_name": staff.first_name,
            "last_name": staff.last_name,
            "rank": staff.rank,
            "role": staff.role,
            "profile_pic": staff.profile_pic.url if staff.profile_pic else None,
            "is_profile_completed": staff.is_profile_completed,
            "email": staff.email,
        })

    if request.method == "PUT":
        payload = json.loads(request.body.decode("utf-8") or "{}")
        for key in payload:
            if hasattr(staff, key):
                setattr(staff, key, payload[key])
        staff.save()
        return JsonResponse({"updated": True})


@csrf_exempt
@require_http_methods(["POST"])
def staff_profile_photo_view(request):
    staff = TrainingStaff.objects.first()
    if not staff:
        return JsonResponse({"message": "No staff found"}, status=404)
    
    uploaded = request.FILES.get("image")
    if not uploaded:
        return JsonResponse({"message": "No image provided"}, status=400)
    try:
        c = compress_uploaded_image(uploaded, max_bytes=IMAGE_MAX_BYTES, max_dim=IMAGE_MAX_DIMENSION)
    except ValueError as exc:
        return JsonResponse({"message": str(exc)}, status=400)
    if c.get("limit_not_met"):
        return JsonResponse({"message": "Unable to meet size limit after compression."}, status=400)
    from django.core.files.base import ContentFile
    content = ContentFile(c["bytes"])
    import time
    ext_map = {"JPEG": ".jpg", "JPG": ".jpg", "PNG": ".png", "WEBP": ".webp"}
    ext = ext_map.get(c["format"].upper(), ".jpg")
    content.name = f"staff_{staff.id}_{int(time.time())}{ext}"
    staff.profile_pic = content
    staff.save()
    return JsonResponse({
        "message": "Photo updated",
        "filePath": staff.profile_pic.url,
        "compression": {"original": c["original_size"], "final": c["size"]},
    })


@require_http_methods(["GET"])
def staff_list_overview_view(request):
    return staff_list_view(request)


@require_http_methods(["GET"])
def staff_analytics_overview_view(request):
    return JsonResponse(
        {
            "totals": {},
            "trends": [],
        }
    )


@csrf_exempt
def staff_collection_view(request):
    if request.method == "GET":
        return staff_list_view(request)
    if request.method == "POST":
        return staff_create_view(request)
    return JsonResponse({"message": "Method not allowed"}, status=405)


@csrf_exempt
@require_http_methods(["PUT", "DELETE"])
def staff_detail_view(request, staff_id):
    if request.method == "PUT":
        return staff_update_view(request, staff_id)
    return staff_delete_view(request, staff_id)


@require_http_methods(["GET"])
def messages_admin_list_view(request):
    return JsonResponse({"messages": []})


@require_http_methods(["GET"])
def messages_my_list_view(request):
    return JsonResponse({"messages": []})


@csrf_exempt
@require_http_methods(["POST"])
def messages_create_view(request):
    return JsonResponse({"created": True})


@csrf_exempt
@require_http_methods(["DELETE"])
def messages_delete_view(request, message_id):
    return JsonResponse({"deleted": True, "id": message_id})


def attendance_events_view(request):
    def event_stream():
        yield "event: heartbeat\n"
        yield 'data: {"status":"ok"}\n\n'

    return StreamingHttpResponse(event_stream(), content_type="text/event-stream")


@csrf_exempt
@require_http_methods(["POST"])
def auth_cadet_login_view(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}
    identifier = (payload.get("identifier") or "").strip()
    if not identifier:
        return JsonResponse({"message": "identifier is required (username or email)"}, status=400)
    try:
        from django.db.models import Q
        cadet = Cadet.objects.filter(Q(username__iexact=identifier) | Q(email__iexact=identifier)).first()
        if not cadet:
            return JsonResponse({"message": "Cadet not found for given identifier"}, status=404)
        token = secrets.token_urlsafe(24)
        return JsonResponse(
            {
                "token": token,
                "role": "cadet",
                "cadetId": cadet.id,
                "staffId": None,
                "isProfileCompleted": bool(cadet.is_profile_completed),
            }
        )
    except Exception as exc:
        return JsonResponse({"message": "Login failed"}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def auth_staff_login_no_pass_view(request):
    return JsonResponse(
        {
            "token": "staff-token",
            "role": "training_staff",
            "cadetId": None,
            "staffId": 1,
            "isProfileCompleted": False,
        }
    )


@require_http_methods(["GET"])
def admin_analytics_overview_view(request):
    return JsonResponse(
        {
            "summary": {},
            "charts": [],
        }
    )


@require_http_methods(["GET"])
def admin_analytics_demographics_view(request):
    return JsonResponse(
        {
            "segments": [],
            "totals": {},
        }
    )


@require_http_methods(["GET"])
def admin_locations_view(request):
    return JsonResponse({"locations": []})


@require_http_methods(["GET"])
def attendance_days_list_view(request):
    return JsonResponse([], safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def attendance_days_create_view(request):
    return JsonResponse({"created": True})


@csrf_exempt
@require_http_methods(["DELETE"])
def attendance_days_delete_view(request, day_id):
    return JsonResponse({"deleted": True, "id": day_id})


@require_http_methods(["GET"])
def attendance_my_history_view(request):
    return JsonResponse([], safe=False)


@require_http_methods(["GET"])
def attendance_my_history_staff_view(request):
    return JsonResponse([], safe=False)


@require_http_methods(["GET"])
def cadet_my_merit_logs_view(request):
    cadet_id = request.GET.get("cadetId")
    if not cadet_id:
        return JsonResponse({"logs": []})
    logs = MeritDemeritLog.objects.filter(cadet_id=cadet_id).order_by("-created_at")
    data = []
    for log in logs:
        data.append(
            {
                "id": log.id,
                "type": log.type,
                "points": log.points,
                "reason": log.reason,
                "issued_by_name": log.issued_by_name,
                "created_at": log.created_at.isoformat(),
            }
        )
    return JsonResponse(data, safe=False)


@require_http_methods(["GET"])
def cadet_my_grades_view(request):
    cadet_id = request.GET.get("cadetId")
    if not cadet_id:
        return JsonResponse(
            {
                "attendanceScore": 0,
                "attendance_present": 0,
                "aptitudeScore": 0,
                "merit_points": 0,
                "demerit_points": 0,
                "subjectScore": 0,
                "prelim_score": 0,
                "midterm_score": 0,
                "final_score": 0,
                "finalGrade": 0,
                "transmutedGrade": "5.00",
                "remarks": "",
            }
        )
    try:
        cadet = Cadet.objects.get(id=cadet_id)
    except Cadet.DoesNotExist:
        return JsonResponse(
            {
                "attendanceScore": 0,
                "attendance_present": 0,
                "aptitudeScore": 0,
                "merit_points": 0,
                "demerit_points": 0,
                "subjectScore": 0,
                "prelim_score": 0,
                "midterm_score": 0,
                "final_score": 0,
                "finalGrade": 0,
                "transmutedGrade": "5.00",
                "remarks": "",
            }
        )
    return JsonResponse(
        {
            "attendanceScore": cadet.attendance_score,
            "attendance_present": cadet.attendance_present,
            "aptitudeScore": cadet.aptitude_score,
            "merit_points": cadet.merit_points,
            "demerit_points": cadet.demerit_points,
            "subjectScore": cadet.subject_score,
            "prelim_score": cadet.prelim_score,
            "midterm_score": cadet.midterm_score,
            "final_score": cadet.final_score,
            "finalGrade": cadet.final_grade,
            "transmutedGrade": cadet.transmuted_grade,
            "remarks": cadet.grade_remarks,
        }
    )


@require_http_methods(["GET"])
def attendance_records_view(request, day_id):
    return JsonResponse([], safe=False)


@require_http_methods(["GET"])
def attendance_records_staff_view(request, day_id):
    return JsonResponse([], safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def attendance_mark_view(request):
    return JsonResponse({"saved": True})


@csrf_exempt
@require_http_methods(["POST"])
def attendance_mark_staff_view(request):
    return JsonResponse({"saved": True})


@csrf_exempt
@require_http_methods(["POST"])
def attendance_import_view(request):
    return JsonResponse(
        {
            "imported": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
        }
    )


@csrf_exempt
@require_http_methods(["POST"])
def integration_rotcmis_validate_view(request):
    return JsonResponse(
        {
            "valid": True,
            "issues": [],
        }
    )


@csrf_exempt
@require_http_methods(["POST"])
def integration_rotcmis_import_view(request):
    return JsonResponse(
        {
            "imported": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
        }
    )


@csrf_exempt
@require_http_methods(["PUT"])
def admin_grades_update_view(request, cadet_id):
    try:
        cadet = Cadet.objects.get(id=cadet_id)
    except Cadet.DoesNotExist:
        return JsonResponse({"message": "Cadet not found"}, status=404)
    try:
        payload = json.loads(request.body.decode() or "{}")
    except json.JSONDecodeError:
        payload = {}
    cadet.prelim_score = float(payload.get("prelimScore", cadet.prelim_score or 0))
    cadet.midterm_score = float(payload.get("midtermScore", cadet.midterm_score or 0))
    cadet.final_score = float(payload.get("finalScore", cadet.final_score or 0))
    cadet.attendance_present = int(
        payload.get("attendancePresent", cadet.attendance_present or 0)
    )
    cadet.merit_points = int(payload.get("meritPoints", cadet.merit_points or 0))
    cadet.demerit_points = int(payload.get("demeritPoints", cadet.demerit_points or 0))
    status = payload.get("status") or ""
    cadet.grade_status = status
    compute_cadet_grades(cadet)
    if status in {"DO", "INC", "T"}:
        cadet.transmuted_grade = "5.00"
        cadet.grade_remarks = "Failed"
    else:
        cadet.grade_remarks = "Passed" if cadet.transmuted_grade != "5.00" else "Failed"
    cadet.save()
    return JsonResponse(cadet_to_dict(cadet))


@require_http_methods(["GET"])
def admin_merit_logs_list_view(request, cadet_id):
    logs = MeritDemeritLog.objects.filter(cadet_id=cadet_id).order_by("-created_at")
    data = []
    for log in logs:
        data.append(
            {
                "id": log.id,
                "type": log.type,
                "points": log.points,
                "reason": log.reason,
                "issued_by_name": log.issued_by_name,
                "created_at": log.created_at.isoformat(),
            }
        )
    return JsonResponse(data, safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def admin_merit_logs_create_view(request):
    try:
        payload = json.loads(request.body.decode() or "{}")
    except json.JSONDecodeError:
        payload = {}
    cadet_id = payload.get("cadetId")
    if not cadet_id:
        return JsonResponse({"message": "cadetId is required"}, status=400)
    try:
        cadet = Cadet.objects.get(id=cadet_id)
    except Cadet.DoesNotExist:
        return JsonResponse({"message": "Cadet not found"}, status=404)
    log = MeritDemeritLog.objects.create(
        cadet=cadet,
        type=payload.get("type") or "merit",
        points=int(payload.get("points", 0)),
        reason=payload.get("reason") or "",
        issued_by_name=payload.get("issuedByName") or "",
    )
    totals = MeritDemeritLog.objects.filter(cadet=cadet).values("type").order_by()
    merit_total = 0
    demerit_total = 0
    for entry in MeritDemeritLog.objects.filter(cadet=cadet):
        if entry.type == "merit":
            merit_total += entry.points
        else:
            demerit_total += entry.points
    cadet.merit_points = merit_total
    cadet.demerit_points = demerit_total
    compute_cadet_grades(cadet)
    if cadet.grade_status in {"DO", "INC", "T"}:
        cadet.transmuted_grade = "5.00"
        cadet.grade_remarks = "Failed"
    else:
        cadet.grade_remarks = "Passed" if cadet.transmuted_grade != "5.00" else "Failed"
    cadet.save()
    return JsonResponse(
        {
            "id": log.id,
            "cadetId": cadet.id,
            "type": log.type,
            "points": log.points,
            "reason": log.reason,
            "issued_by_name": log.issued_by_name,
        }
    )


@csrf_exempt
@require_http_methods(["DELETE"])
def admin_merit_logs_delete_view(request, log_id):
    try:
        log = MeritDemeritLog.objects.get(id=log_id)
    except MeritDemeritLog.DoesNotExist:
        return JsonResponse({"message": "Log not found"}, status=404)
    cadet = log.cadet
    log.delete()
    merit_total = 0
    demerit_total = 0
    for entry in MeritDemeritLog.objects.filter(cadet=cadet):
        if entry.type == "merit":
            merit_total += entry.points
        else:
            demerit_total += entry.points
    cadet.merit_points = merit_total
    cadet.demerit_points = demerit_total
    compute_cadet_grades(cadet)
    if cadet.grade_status in {"DO", "INC", "T"}:
        cadet.transmuted_grade = "5.00"
        cadet.grade_remarks = "Failed"
    else:
        cadet.grade_remarks = "Passed" if cadet.transmuted_grade != "5.00" else "Failed"
    cadet.save()
    return JsonResponse({"deleted": True, "id": log_id})


@require_http_methods(["GET"])
def attendance_cadet_history_view(request, cadet_id):
    return JsonResponse([], safe=False)


@csrf_exempt
@require_http_methods(["POST"])
def admin_sync_lifetime_merits_view(request):
    cadets = Cadet.objects.all()
    synced = 0
    for cadet in cadets:
        merit_total = 0
        demerit_total = 0
        for entry in MeritDemeritLog.objects.filter(cadet=cadet):
            if entry.type == "merit":
                merit_total += entry.points
            else:
                demerit_total += entry.points
        cadet.merit_points = merit_total
        cadet.demerit_points = demerit_total
        compute_cadet_grades(cadet)
        if cadet.grade_status in {"DO", "INC", "T"}:
            cadet.transmuted_grade = "5.00"
            cadet.grade_remarks = "Failed"
        else:
            cadet.grade_remarks = (
                "Passed" if cadet.transmuted_grade != "5.00" else "Failed"
            )
        cadet.save()
        synced += 1
    return JsonResponse(
        {
            "synced": True,
            "syncedCount": synced,
            "totalCadets": cadets.count(),
        }
    )


@csrf_exempt
@require_http_methods(["POST"])
def integration_grades_import_view(request):
    return JsonResponse({"imported": 0, "updated": 0, "skipped": 0, "errors": []})


@csrf_exempt
@require_http_methods(["POST"])
def integration_ledger_import_view(request):
    return JsonResponse({"imported": 0, "updated": 0, "skipped": 0, "errors": []})


@csrf_exempt
@require_http_methods(["POST"])
def admin_import_staff_view(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        return JsonResponse({"message": "No file uploaded"}, status=400)
    try:
        rows = load_rows_from_upload(uploaded)
    except ValueError as exc:
        return JsonResponse({"message": str(exc)}, status=400)

    imported = 0
    updated = 0
    skipped = 0
    errors = []

    for row in rows:
        try:
            email = find_column_value(row, ["Email", "email", "E-mail"])
            first_name = find_column_value(row, ["First Name", "first_name", "FName"])
            last_name = find_column_value(row, ["Last Name", "last_name", "LName"])
            middle_name = (
                find_column_value(row, ["Middle Name", "middle_name", "MName"]) or ""
            )

            if not first_name or not last_name:
                full_name = find_column_value(
                    row, ["Name", "name", "Full Name", "Staff Name"]
                )
                if full_name:
                    parts = str(full_name).split(" ")
                    if len(parts) >= 2:
                        first_name = " ".join(parts[:-1])
                        last_name = parts[-1]

            if not first_name or not last_name:
                skipped += 1
                errors.append("Missing staff first/last name")
                continue

            rank = find_column_value(row, ["Rank", "rank"]) or ""
            contact_number = find_column_value(
                row, ["Contact Number", "Phone", "contact_number"]
            ) or ""
            role = find_column_value(row, ["Role", "role", "Position"]) or "Instructor"

            staff_defaults = {
                "rank": rank,
                "first_name": first_name,
                "middle_name": middle_name,
                "last_name": last_name,
                "suffix_name": "",
                "email": email or "",
                "contact_number": contact_number,
                "role": role,
            }

            lookup_email = staff_defaults["email"] or None

            if lookup_email:
                staff, created = TrainingStaff.objects.get_or_create(
                    email=lookup_email, defaults=staff_defaults
                )
            else:
                staff, created = TrainingStaff.objects.get_or_create(
                    first_name=first_name,
                    last_name=last_name,
                    rank=rank,
                    defaults=staff_defaults,
                )

            if not created:
                for key, value in staff_defaults.items():
                    setattr(staff, key, value)
                staff.save()
                updated += 1
            else:
                imported += 1
        except Exception as exc:
            skipped += 1
            errors.append(str(exc))

    return JsonResponse(
        {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10],
            "message": f"Import complete. Success: {imported + updated}, Failed: {skipped}",
        }
    )


@csrf_exempt
@require_http_methods(["POST"])
def staff_create_view(request):
    if request.content_type and "application/json" in request.content_type:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    else:
        payload = request.POST
    staff = TrainingStaff.objects.create(
        rank=payload.get("rank", ""),
        first_name=payload.get("first_name", ""),
        middle_name=payload.get("middle_name", ""),
        last_name=payload.get("last_name", ""),
        suffix_name=payload.get("suffix_name", ""),
        email=payload.get("email", ""),
        contact_number=payload.get("contact_number", ""),
        role=payload.get("role", "Instructor"),
    )
    return JsonResponse({"created": True, "id": staff.id})


@csrf_exempt
@require_http_methods(["PUT"])
def staff_update_view(request, staff_id):
    try:
        staff = TrainingStaff.objects.get(id=staff_id)
    except TrainingStaff.DoesNotExist:
        return JsonResponse({"message": "Staff not found"}, status=404)
    if request.content_type and "application/json" in request.content_type:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    else:
        payload = request.POST
    staff.rank = payload.get("rank", staff.rank)
    staff.first_name = payload.get("first_name", staff.first_name)
    staff.middle_name = payload.get("middle_name", staff.middle_name)
    staff.last_name = payload.get("last_name", staff.last_name)
    staff.suffix_name = payload.get("suffix_name", staff.suffix_name)
    staff.email = payload.get("email", staff.email)
    staff.contact_number = payload.get("contact_number", staff.contact_number)
    staff.role = payload.get("role", staff.role)
    staff.save()
    return JsonResponse({"updated": True, "id": staff.id})


@csrf_exempt
@require_http_methods(["DELETE"])
def staff_delete_view(request, staff_id):
    try:
        staff = TrainingStaff.objects.get(id=staff_id)
    except TrainingStaff.DoesNotExist:
        return JsonResponse({"message": "Staff not found"}, status=404)
    staff.is_archived = True
    staff.save()
    return JsonResponse({"deleted": True, "id": staff_id})


@csrf_exempt
@require_http_methods(["POST"])
def auth_location_view(request):
    return JsonResponse({"saved": True})


@require_http_methods(["GET"])
def admin_cadets_export_completed_view(request):
    return JsonResponse({"message": "no completed cadets", "has_data": False}, status=404)


@csrf_exempt
@require_http_methods(["DELETE"])
def admin_cadets_prune_completed_view(request):
    return JsonResponse({"message": "No completed cadets to delete."})


@require_http_methods(["GET"])
def admin_export_view(request, export_type):
    return JsonResponse(
        {
            "message": "export generated",
            "type": export_type,
        }
    )


@require_http_methods(["GET"])
def image_admin_view(request, admin_id):
    try:
        admin_profile = AdminProfile.objects.get(id=admin_id)
        if admin_profile.profile_pic:
            from django.shortcuts import redirect
            return redirect(admin_profile.profile_pic.url)
    except AdminProfile.DoesNotExist:
        pass
    return transparent_png_response()


@require_http_methods(["GET"])
def image_staff_view(request, staff_id):
    try:
        staff = TrainingStaff.objects.get(id=staff_id)
        if staff.profile_pic:
            from django.shortcuts import redirect
            return redirect(staff.profile_pic.url)
    except TrainingStaff.DoesNotExist:
        pass
    return transparent_png_response()


@require_http_methods(["GET"])
def image_cadet_view(request, cadet_id):
    try:
        cadet = Cadet.objects.get(id=cadet_id)
        if cadet.profile_pic:
            from django.shortcuts import redirect
            return redirect(cadet.profile_pic.url)
    except Cadet.DoesNotExist:
        pass
    return transparent_png_response()


@csrf_exempt
@require_http_methods(["GET", "POST"])
def excuse_collection_view(request):
    meta_path = settings.BASE_DIR / "excuses_meta.json"
    if request.method == "GET":
        try:
            if meta_path.exists():
                raw = meta_path.read_text(encoding="utf-8")
                data = json.loads(raw or "[]")
                return JsonResponse(data, safe=False)
        except Exception as exc:
            logging.error("Failed to read excuses meta: %s", exc)
        return JsonResponse([], safe=False)

    uploaded = request.FILES.get("file")
    date_absent = request.POST.get("date_absent")
    reason = request.POST.get("reason")
    cadet_id = request.POST.get("cadetId")
    if not uploaded or not date_absent or not reason:
        return JsonResponse({"message": "Missing fields"}, status=400)

    content_type = uploaded.content_type or ""
    original_name = uploaded.name
    compression_info = None

    try:
        if content_type.startswith("image/"):
            c = compress_uploaded_image(uploaded, max_bytes=IMAGE_MAX_BYTES, max_dim=IMAGE_MAX_DIMENSION)
            if c.get("limit_not_met"):
                return JsonResponse({"message": "Unable to meet size limit after compression."}, status=400)
            content = ContentFile(c["bytes"])
            ext_map = {"JPEG": ".jpg", "JPG": ".jpg", "PNG": ".png", "WEBP": ".webp"}
            ext = ext_map.get(c["format"].upper(), os.path.splitext(original_name)[1] or ".jpg")
            filename = f"excuses/excuse_{int(datetime.utcnow().timestamp())}{ext}"
            compression_info = {"type": "image", "original": c["original_size"], "final": c["size"]}
        elif content_type == "application/pdf":
            buf = uploaded.read()
            c = compress_pdf_bytes(buf, max_bytes=PDF_MAX_BYTES)
            bytes_to_save = c["bytes"]
            if c.get("limit_not_met"):
                logging.warning("PDF still above limit after compression: %s -> %s", c["original_size"], c["size"])
            content = ContentFile(bytes_to_save)
            filename = f"excuses/excuse_{int(datetime.utcnow().timestamp())}.pdf"
            compression_info = {"type": "pdf", "original": c["original_size"], "final": c["size"], "skipped": c.get("skipped", False)}
        else:
            content = uploaded
            filename = f"excuses/{int(datetime.utcnow().timestamp())}_{original_name}"

        saved_path = default_storage.save(filename, content)
        file_url = default_storage.url(saved_path)

        item = {
            "date_absent": date_absent,
            "reason": reason,
            "status": "pending",
            "file_url": file_url,
            "original_name": original_name,
            "uploaded_at": datetime.utcnow().isoformat() + "Z",
        }
        if cadet_id:
            item["cadet_id"] = int(cadet_id)
        try:
            existing = []
            if meta_path.exists():
                existing = json.loads(meta_path.read_text(encoding="utf-8") or "[]")
            existing.insert(0, item)
            meta_path.write_text(json.dumps(existing[:200]), encoding="utf-8")
        except Exception as exc:
            logging.error("Failed to write excuses meta: %s", exc)

        return JsonResponse({"message": "Submitted", "item": item, "compression": compression_info})
    except Exception as exc:
        logging.exception("Excuse upload failed: %s", exc)
        return JsonResponse({"message": "Upload failed"}, status=500)


@require_http_methods(["GET"])
def absence_analytics_view(request):
    try:
        from collections import defaultdict
        import math
        start = request.GET.get("startDate")
        end = request.GET.get("endDate")
        battalion = request.GET.get("battalion")
        company = request.GET.get("company")
        platoon = request.GET.get("platoon")
        year_level = request.GET.get("yearLevel")
        type_filter = request.GET.get("type")  # excused/unexcused
        severity_filter = request.GET.get("severity")  # low/medium/high

        cadets = Cadet.objects.all()
        if battalion:
            cadets = cadets.filter(battalion=battalion)
        if company:
            cadets = cadets.filter(company=company)
        if platoon:
            cadets = cadets.filter(platoon=platoon)
        if year_level:
            cadets = cadets.filter(year_level=year_level)

        cadet_map = {c.id: c for c in cadets}
        meta_path = settings.BASE_DIR / "excuses_meta.json"
        excuses = []
        if meta_path.exists():
            try:
                excuses = json.loads(meta_path.read_text(encoding="utf-8") or "[]")
            except Exception:
                excuses = []

        def within_range(d):
            if (not start) and (not end):
                return True
            try:
                from datetime import date
                dt = datetime.strptime(d, "%Y-%m-%d").date()
                if start:
                    s = datetime.strptime(start, "%Y-%m-%d").date()
                    if dt < s:
                        return False
                if end:
                    e = datetime.strptime(end, "%Y-%m-%d").date()
                    if dt > e:
                        return False
            except Exception:
                return True
            return True

        def classify(exc):
            reason = (exc.get("reason") or "").lower()
            status = (exc.get("status") or "").lower()
            etype = "excused" if status == "approved" else "unexcused"
            if "medical" in reason or len(reason) > 120:
                severity = "high"
            elif "official" in reason or "event" in reason:
                severity = "low"
            else:
                severity = "medium"
            return etype, severity

        per_cadet_excuses = defaultdict(list)
        timeline_counts = defaultdict(int)
        for exc in excuses:
            cadet_id = exc.get("cadet_id")
            date_str = exc.get("date_absent")
            if not cadet_id or cadet_id not in cadet_map or not date_str:
                continue
            if not within_range(date_str):
                continue
            etype, severity = classify(exc)
            if (type_filter and etype != type_filter) or (severity_filter and severity != severity_filter):
                continue
            per_cadet_excuses[cadet_id].append({"date": date_str, "type": etype, "severity": severity})
            timeline_counts[date_str] += 1

        cadet_rows = []
        group_by = {
            "battalion": defaultdict(lambda: {"absences": 0, "cadets": 0}),
            "company": defaultdict(lambda: {"absences": 0, "cadets": 0}),
            "platoon": defaultdict(lambda: {"absences": 0, "cadets": 0}),
            "year_level": defaultdict(lambda: {"absences": 0, "cadets": 0}),
        }
        for cid, cadet in cadet_map.items():
            present = cadet.attendance_present or 0
            total = cadet.attendance_total or 0
            total_absences = max(0, total - present)
            excuse_count = len(per_cadet_excuses.get(cid, []))
            percent = (total_absences / total) * 100 if total > 0 else 0.0
            # Simple trend: last 30 days vs previous 30 days from excuses
            today = datetime.utcnow().date()
            last30 = 0
            prev30 = 0
            for e in per_cadet_excuses.get(cid, []):
                try:
                    d = datetime.strptime(e["date"], "%Y-%m-%d").date()
                    delta = (today - d).days
                    if 0 <= delta <= 30:
                        last30 += 1
                    elif 31 <= delta <= 60:
                        prev30 += 1
                except Exception:
                    continue
            trend = last30 - prev30
            risk_score = round(0.6 * percent + 4.0 * max(0, trend), 2)
            cadet_rows.append(
                {
                    "cadetId": cid,
                    "name": f"{cadet.last_name}, {cadet.first_name}",
                    "battalion": cadet.battalion,
                    "company": cadet.company,
                    "platoon": cadet.platoon,
                    "year_level": cadet.year_level,
                    "present": present,
                    "total": total,
                    "total_absences": total_absences,
                    "excuse_events": excuse_count,
                    "percent": round(percent, 2),
                    "trend": trend,
                    "risk_score": risk_score,
                }
            )
            group_by["battalion"][cadet.battalion]["absences"] += total_absences
            group_by["battalion"][cadet.battalion]["cadets"] += 1
            group_by["company"][cadet.company]["absences"] += total_absences
            group_by["company"][cadet.company]["cadets"] += 1
            group_by["platoon"][cadet.platoon]["absences"] += total_absences
            group_by["platoon"][cadet.platoon]["cadets"] += 1
            group_by["year_level"][cadet.year_level]["absences"] += total_absences
            group_by["year_level"][cadet.year_level]["cadets"] += 1

        cadet_rows.sort(key=lambda x: (-x["risk_score"], -x["percent"], -x["total_absences"]))
        for idx, row in enumerate(cadet_rows, start=1):
            row["rank"] = idx

        def to_series(dct):
            out = []
            for k, v in dct.items():
                if v["cadets"] > 0:
                    rate = v["absences"] / v["cadets"]
                else:
                    rate = 0.0
                out.append({"key": k or "N/A", "avg_absences": round(rate, 2)})
            out.sort(key=lambda x: -x["avg_absences"])
            return out

        timeline = [{"date": k, "absences": v} for k, v in timeline_counts.items()]
        timeline.sort(key=lambda x: x["date"])

        return JsonResponse(
            {
                "cadets": cadet_rows,
                "groups": {
                    "battalion": to_series(group_by["battalion"]),
                    "company": to_series(group_by["company"]),
                    "platoon": to_series(group_by["platoon"]),
                    "year_level": to_series(group_by["year_level"]),
                },
                "timeline": timeline,
                "at_risk": cadet_rows[:10],
                "filters": {
                    "applied": {
                        "startDate": start,
                        "endDate": end,
                        "battalion": battalion,
                        "company": company,
                        "platoon": platoon,
                        "yearLevel": year_level,
                        "type": type_filter,
                        "severity": severity_filter,
                    }
                },
            },
            safe=False,
        )
    except Exception as exc:
        logging.exception("Analytics failed: %s", exc)
        return JsonResponse({"message": "Analytics computation failed"}, status=500)


@require_http_methods(["GET"])
def media_serve_view(request, path):
    try:
        allowed_prefixes = ("admin_profiles/", "staff_profiles/", "cadet_profiles/", "excuses/")
        if not path.startswith(allowed_prefixes):
            return JsonResponse({"message": "Forbidden"}, status=403)
        if not default_storage.exists(path):
            return JsonResponse({"message": "Not Found"}, status=404)
        f = default_storage.open(path, "rb")
        data = f.read()
        f.close()
        ctype, _ = mimetypes.guess_type(path)
        resp = HttpResponse(data, content_type=ctype or "application/octet-stream")
        resp["Cache-Control"] = "max-age=3600, public"
        return resp
    except Exception as exc:
        logging.exception("Media serve failed for %s: %s", path, exc)
        return JsonResponse({"message": "Error serving media"}, status=500)

