import csv
import json
import os
from typing import Iterable, Set, Tuple, List

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from rgms.models import Cadet

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def normalize_header(s: str) -> str:
    return (s or "").strip().lower().replace(" ", "").replace("_", "")


def normalize_value(s: str) -> str:
    return (s or "").strip()


def detect_student_id_header(headers: Iterable[str]) -> str:
    candidates = {normalize_header(h): h for h in headers}
    for key in ["studentid", "idnumber", "student_no", "studentnumber", "sid"]:
        if key in candidates:
            return candidates[key]
    for raw in headers:
        if raw.lower() in {"student id", "student_id", "id number"}:
            return raw
    return list(headers)[0]


def read_ids_from_csv(path: str) -> Set[str]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        if not headers:
            raise CommandError("CSV has no headers")
        sid_col = detect_student_id_header(headers)
        ids: Set[str] = set()
        for row in reader:
            val = normalize_value(row.get(sid_col, ""))
            if val:
                ids.add(val)
        return ids


def read_ids_from_xlsx(path: str) -> Set[str]:
    if load_workbook is None:
        raise CommandError("openpyxl is not installed; cannot read .xlsx files")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise CommandError("XLSX contains no rows")
    headers = [str(h or "").strip() for h in rows[0]]
    sid_col_name = detect_student_id_header(headers)
    try:
        idx = headers.index(sid_col_name)
    except ValueError:
        idx = 0
    ids: Set[str] = set()
    for r in rows[1:]:
        if not r:
            continue
        val = r[idx]
        if val is None:
            continue
        ids.add(str(val).strip())
    return ids


class Command(BaseCommand):
    help = "Delete cadet records that are NOT present in a newly imported dataset (by student_id)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to CSV/XLSX import file")
        parser.add_argument(
            "--apply",
            action="store_true",
            help="Apply deletions. Omit for a dry-run summary.",
        )
        parser.add_argument(
            "--backup",
            default="deleted_cadets_backup.json",
            help="Path to write JSON backup of deleted cadets (when --apply).",
        )

    def handle(self, *args, **options):
        path = options["file"]
        apply = options["apply"]
        backup_path = options["backup"]

        if not os.path.exists(path):
            raise CommandError(f"File not found: {path}")

        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            import_ids = read_ids_from_csv(path)
        elif ext in {".xlsx", ".xlsm"}:
            import_ids = read_ids_from_xlsx(path)
        else:
            raise CommandError("Unsupported file type. Use .csv or .xlsx")

        if not import_ids:
            raise CommandError("No student IDs found in import file")

        existing_ids = set(Cadet.objects.values_list("student_id", flat=True))
        preserve = {normalize_value(sid) for sid in import_ids}
        to_delete_ids = [sid for sid in existing_ids if normalize_value(sid) not in preserve]

        self.stdout.write(self.style.NOTICE(f"Import IDs: {len(import_ids)}"))
        self.stdout.write(self.style.NOTICE(f"Existing cadets: {len(existing_ids)}"))
        self.stdout.write(self.style.WARNING(f"Cadets to delete: {len(to_delete_ids)}"))

        if not apply:
            self.stdout.write(self.style.SUCCESS("Dry run complete. No records were deleted. Use --apply to perform deletion."))
            return

        backup_rows: List[dict] = list(
            Cadet.objects.filter(student_id__in=to_delete_ids).values(
                "id",
                "student_id",
                "first_name",
                "last_name",
                "email",
                "username",
                "company",
                "platoon",
                "cadet_course",
                "status",
            )
        )
        with open(backup_path, "w", encoding="utf-8") as bf:
            json.dump({"deleted": backup_rows}, bf, ensure_ascii=False, indent=2)

        with transaction.atomic():
            deleted_count, _ = Cadet.objects.filter(student_id__in=to_delete_ids).delete()

        self.stdout.write(self.style.SUCCESS(f"Deleted {deleted_count} cadet record(s). Backup saved to: {backup_path}"))
