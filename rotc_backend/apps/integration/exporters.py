"""
Export utilities for generating CSV and Excel files
"""
import csv
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


class CSVExporter:
    """Utility class for exporting data to CSV format"""
    
    @staticmethod
    def export_to_response(data: List[Dict[str, Any]], filename: str, headers: Optional[List[str]] = None) -> HttpResponse:
        """
        Export data to CSV and return as HTTP response
        
        Args:
            data: List of dictionaries containing the data to export
            filename: Name of the CSV file
            headers: Optional list of header names. If None, uses keys from first data item
            
        Returns:
            HttpResponse with CSV content
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        if not data:
            return response
        
        # Use provided headers or extract from first item
        if headers is None:
            headers = list(data[0].keys())
        
        writer = csv.DictWriter(response, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    @staticmethod
    def export_to_string(data: List[Dict[str, Any]], headers: Optional[List[str]] = None) -> str:
        """
        Export data to CSV string
        
        Args:
            data: List of dictionaries containing the data to export
            headers: Optional list of header names
            
        Returns:
            CSV content as string
        """
        if not data:
            return ""
        
        output = io.StringIO()
        
        if headers is None:
            headers = list(data[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue()


class ExcelExporter:
    """Utility class for exporting data to Excel format"""
    
    # Style constants
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def create_workbook(data: List[Dict[str, Any]], sheet_name: str = "Sheet1", 
                       headers: Optional[List[str]] = None) -> Workbook:
        """
        Create an Excel workbook from data
        
        Args:
            data: List of dictionaries containing the data to export
            sheet_name: Name of the worksheet
            headers: Optional list of header names
            
        Returns:
            Workbook object
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return wb
        
        # Use provided headers or extract from first item
        if headers is None:
            headers = list(data[0].keys())
        
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = ExcelExporter.HEADER_FONT
            cell.fill = ExcelExporter.HEADER_FILL
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
        
        # Write data rows
        for row_num, item in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = item.get(header, "")
                # Convert datetime objects to strings
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    @staticmethod
    def export_to_response(data: List[Dict[str, Any]], filename: str, 
                          sheet_name: str = "Sheet1", headers: Optional[List[str]] = None) -> HttpResponse:
        """
        Export data to Excel and return as HTTP response
        
        Args:
            data: List of dictionaries containing the data to export
            filename: Name of the Excel file
            sheet_name: Name of the worksheet
            headers: Optional list of header names
            
        Returns:
            HttpResponse with Excel content
        """
        wb = ExcelExporter.create_workbook(data, sheet_name, headers)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def create_multi_sheet_workbook(sheets: Dict[str, List[Dict[str, Any]]]) -> Workbook:
        """
        Create an Excel workbook with multiple sheets
        
        Args:
            sheets: Dictionary mapping sheet names to data lists
            
        Returns:
            Workbook object
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            
            if not data:
                continue
            
            headers = list(data[0].keys())
            
            # Write headers with styling
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = ExcelExporter.HEADER_FONT
                cell.fill = ExcelExporter.HEADER_FILL
                cell.alignment = ExcelExporter.HEADER_ALIGNMENT
            
            # Write data rows
            for row_num, item in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(header))
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
